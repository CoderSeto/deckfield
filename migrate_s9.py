"""
Migrate real Season 9 data from Baccer_Game_S9_Stats.xlsm into deckfield.db.

Known limitation: the per-round EX Bonus Score (B) isn't recoverable from
the workbook -- only the current cumulative total is visible (Rankings
column B). Migrated games are inserted with ex_bonus=0, so computed TOT
(and anything downstream of it: Cups, OVR) will read LOWER than the real
Rankings tab by roughly that team's cumulative B. RW and LW don't depend
on B and should match closely.
"""

import re
import openpyxl
from collections import defaultdict
from deckfield_ratings import init_db, seed_region_distances, add_team, set_team_season, add_game

XLSM_PATH = "/mnt/user-data/uploads/Baccer_Game_S9_Stats.xlsm"
SEASON = 9

LEAGUE_DIVISION_MAP = {
    "First": 1, "Second": 2, "Third": 3, "Fourth": 4, "Fifth": 5,
    "Sixth": 6, "Seventh": 7, "Eighth": 8, "Ninth": 9, "Tenth": 10,
}
REGION_NAME_FIX = {"Lily Valley": "LilyValley"}  # matches region_distances keys

GAME_PATTERN = re.compile(
    r'^(\d)([RLSPF])!(-?\d+)\+(-?\d+)-(-?\d+)\^(-?[\d.]+)\((-?[\d.]+)\)'
    r'\[(-?[\d.]+)\]\{(-?[\d.]+)\}(-?[\d.]+)$'
)
WALKOVER_PATTERN = re.compile(r'^(\d)([RLSPF])-$')
EX_ARCHIVE_PATTERN = re.compile(r'/(-?[\d.]+)/')
GAME_COLS = list(range(12, 97))  # Calcs columns L..CR, in play order
ARCHIVE_LAST_ROW = 785  # rows 2..785 are real games; beyond that is unplayed template

FATIGUE_COL_START = 960  # Calcs column AJX -- position 1 (R1) is the starting baseline,
FATIGUE_COLS = list(range(FATIGUE_COL_START, FATIGUE_COL_START + len(GAME_COLS)))
# positions 2+ are per-round deltas, aligned 1:1 with GAME_COLS by position (verified:
# R/L labels match exactly at every position; only cup-slot labels differ, S# vs D#/P#)


def migrate_fatigue(wb, conn):
    """Starting fatigue (position 1) goes into team_seasons; every
    subsequent position is that round's delta, stored in fatigue_deltas."""
    ws = wb["Calcs"]
    updated_seasons, delta_count = 0, 0
    for row in range(3, 163):
        dex = ws.cell(row=row, column=1).value
        if dex is None:
            continue
        start_val = ws.cell(row=row, column=FATIGUE_COLS[0]).value
        if start_val is not None:
            conn.execute(
                "UPDATE team_seasons SET starting_fatigue = ? WHERE team_id = ? AND season = ?",
                (float(start_val), dex, SEASON),
            )
            updated_seasons += 1
        for idx, col in enumerate(FATIGUE_COLS[1:], start=2):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            conn.execute(
                "INSERT OR REPLACE INTO fatigue_deltas (team_id, season, round, delta) VALUES (?,?,?,?)",
                (dex, SEASON, idx, float(val)),
            )
            delta_count += 1
    conn.commit()
    return updated_seasons, delta_count


def migrate_ex_bonus(wb, conn):
    """
    Backfill the real per-round EX Bonus Score from the Archive tab.
    Archive column I names the winner, column P holds their full packed
    game string, column S holds the bonus as '/n/'. Only the winner gets
    a bonus. Matched back to the already-migrated games table by team
    pair + scores + Elo (a very high-uniqueness compound key).
    """
    ws = wb["Archive"]
    team_ids_by_name = {r["name"]: r["team_id"] for r in
                         conn.execute("SELECT team_id, name FROM teams").fetchall()}

    matched, unmatched = 0, []
    for row in range(2, ARCHIVE_LAST_ROW + 1):
        winner_name = ws.cell(row=row, column=9).value
        winner_str = ws.cell(row=row, column=16).value
        ex_cell = ws.cell(row=row, column=19).value
        if winner_name is None or winner_str is None:
            continue
        m = GAME_PATTERN.match(str(winner_str))
        ex_m = EX_ARCHIVE_PATTERN.match(str(ex_cell)) if ex_cell else None
        if not m or not ex_m:
            unmatched.append((row, winner_name, winner_str, ex_cell))
            continue

        _, gtype, opp_dex, pf, pa, *_ , elo_after = m.groups()
        opp_dex, pf, pa, elo_after = int(opp_dex), int(pf), int(pa), float(elo_after)
        ex_value = float(ex_m.group(1))
        winner_id = team_ids_by_name.get(winner_name)
        if winner_id is None:
            unmatched.append((row, winner_name, "no team_id match", None))
            continue

        row_a = conn.execute("""
            SELECT game_id FROM games WHERE season = 9 AND team_a = ? AND team_b = ?
            AND pf_a = ? AND pa_a = ? AND elo_a_after = ?
        """, (winner_id, opp_dex, pf, pa, elo_after)).fetchone()
        if row_a:
            conn.execute("UPDATE games SET ex_bonus_a = ? WHERE game_id = ?", (ex_value, row_a["game_id"]))
            matched += 1
            continue
        row_b = conn.execute("""
            SELECT game_id FROM games WHERE season = 9 AND team_b = ? AND team_a = ?
            AND pa_a = ? AND pf_a = ? AND elo_b_after = ?
        """, (winner_id, opp_dex, pf, pa, elo_after)).fetchone()
        if row_b:
            conn.execute("UPDATE games SET ex_bonus_b = ? WHERE game_id = ?", (ex_value, row_b["game_id"]))
            matched += 1
            continue
        unmatched.append((row, winner_name, winner_str, ex_cell))

    conn.commit()
    return matched, unmatched


def migrate_teams(wb):
    ws = wb["Rankings"]
    count = 0
    for row in range(2, 162):
        dex = ws.cell(row=row, column=1).value
        if dex is None:
            continue
        name = ws.cell(row=row, column=3).value
        region = ws.cell(row=row, column=2).value
        region = REGION_NAME_FIX.get(region, region)
        primary_type = ws.cell(row=row, column=77).value
        accolades = ws.cell(row=row, column=79).value or None
        basclm = ws.cell(row=row, column=69).value
        leag_text = ws.cell(row=row, column=74).value
        league_division = LEAGUE_DIVISION_MAP[leag_text]
        ex = ws.cell(row=row, column=55).value

        add_team(dex, name, region, primary_type=str(primary_type), accolades=accolades)
        set_team_season(dex, SEASON, league_division=league_division, basclm=basclm, ex=ex)
        count += 1
    return count


def parse_game_records(wb):
    ws = wb["Calcs"]
    records = []
    walkovers = []
    unmatched = []
    for row in range(3, 163):
        dex = ws.cell(row=row, column=1).value
        if dex is None:
            continue
        for idx, col in enumerate(GAME_COLS, start=1):
            v = ws.cell(row=row, column=col).value
            if v is None or v == "-":
                continue
            v = str(v)
            m = GAME_PATTERN.match(v)
            if m:
                result, gtype, opp, pf, pa, raw, spread, interest, dscr, elo = m.groups()
                records.append({
                    "dex": dex, "round": idx, "game_type": gtype, "opp": int(opp),
                    "pf": int(pf), "pa": int(pa), "result": int(result),
                    "raw": float(raw), "spread": float(spread), "interest": float(interest),
                    "dscr": float(dscr), "elo": float(elo),
                })
                continue
            wm = WALKOVER_PATTERN.match(v)
            if wm:
                result, gtype = wm.groups()
                walkovers.append({"dex": dex, "round": idx, "game_type": gtype, "result": int(result)})
                continue
            unmatched.append((row, col, v))
    return records, walkovers, unmatched


def dedupe_games(records):
    by_key = defaultdict(list)
    for r in records:
        key = (r["round"], frozenset({r["dex"], r["opp"]}))
        by_key[key].append(r)

    games, skipped = [], 0
    for key, recs in by_key.items():
        if len(recs) != 2:
            skipped += 1
            continue
        a, b = recs
        games.append({
            "round": a["round"], "game_type": a["game_type"],
            "team_a": a["dex"], "team_b": b["dex"],
            "pf_a": a["pf"], "pa_a": a["pa"], "result_a": a["result"],
            "raw_goal_score_a": a["raw"], "raw_goal_score_b": b["raw"],
            "spread_a": a["spread"], "interest_score": a["interest"],
            "dscr_a": a["dscr"], "dscr_b": b["dscr"],
            "elo_a_after": a["elo"], "elo_b_after": b["elo"],
        })
    return games, skipped


def migrate_games(games):
    for g in sorted(games, key=lambda g: g["round"]):
        add_game(
            SEASON, g["round"], g["game_type"], g["team_a"], g["team_b"],
            pf_a=g["pf_a"], pa_a=g["pa_a"], result_a=g["result_a"],
            raw_goal_score_a=g["raw_goal_score_a"], raw_goal_score_b=g["raw_goal_score_b"],
            spread_a=g["spread_a"], interest_score=g["interest_score"],
            dscr_a=g["dscr_a"], dscr_b=g["dscr_b"], ex_bonus_a=0, ex_bonus_b=0,
            elo_a_after=g["elo_a_after"], elo_b_after=g["elo_b_after"],
        )


def tag_rds_cup_games(wb, conn):
    """
    Tag real RDS Cup games (currently generic game_type='S') with which cup,
    bracket, and cup-round they belong to, using the confirmed Archive
    row-blocks for round 1 and round 2 of both Draw and Process. Matched by
    team pair + the already-validated absolute round number for each block
    (round1: abs round 4=Draw, 5=Process; round2: abs round 10=Draw, 11=Process).
    """
    ws_archive = wb["Archive"]

    # team -> cup, from each cup tab's real seed list (col A=seed, col B=team)
    team_to_cup = {}
    for cup in ("Ribbon", "Dream", "Star"):
        ws = wb[cup]
        for row in range(1, 65):
            team = ws.cell(row=row, column=2).value
            if team and team != "bye":
                team_to_cup[team] = cup

    blocks = [
        (242, 305, 4, "Draw", 1),
        (306, 369, 5, "Process", 1),
        (690, 737, 10, "Draw", 2),
        (738, 785, 11, "Process", 2),
    ]

    tagged, unmatched = 0, []
    for start, end, abs_round, bracket, cup_round in blocks:
        for row in range(start, end + 1):
            winner = ws_archive.cell(row=row, column=9).value
            loser = ws_archive.cell(row=row, column=12).value
            if not winner or not loser:
                continue
            cup_name = team_to_cup.get(winner) or team_to_cup.get(loser)
            if cup_name is None:
                unmatched.append((row, winner, loser))
                continue

            winner_id = conn.execute("SELECT team_id FROM teams WHERE name=?", (winner,)).fetchone()
            loser_id = conn.execute("SELECT team_id FROM teams WHERE name=?", (loser,)).fetchone()
            if winner_id is None or loser_id is None:
                unmatched.append((row, winner, loser))
                continue
            w, l = winner_id["team_id"], loser_id["team_id"]

            cur = conn.execute("""
                UPDATE games SET cup_name=?, cup_bracket=?, cup_round=?
                WHERE season=9 AND round=? AND game_type='S'
                AND ((team_a=? AND team_b=?) OR (team_a=? AND team_b=?))
            """, (cup_name, bracket, cup_round, abs_round, w, l, l, w))
            if cur.rowcount == 0:
                unmatched.append((row, winner, loser))
            else:
                tagged += cur.rowcount

    conn.commit()
    return tagged, unmatched


def migrate_host_region(wb, conn):
    """
    Backfill host_region for historical games from the Archive tab.
    Column I names the winner, column G is 'H' or 'A' for whether the
    WINNER was home or away. If the winner was away, the loser hosted.
    Matched back to the games table the same way as the EX bonus backfill:
    by team pair, scores, and Elo.
    """
    ws = wb["Archive"]
    team_ids_by_name = {r["name"]: r["team_id"] for r in
                         conn.execute("SELECT team_id, name FROM teams").fetchall()}
    team_regions = {r["team_id"]: r["region"] for r in
                     conn.execute("SELECT team_id, region FROM teams").fetchall()}

    matched, unmatched = 0, []
    for row in range(2, ARCHIVE_LAST_ROW + 1):
        winner_name = ws.cell(row=row, column=9).value
        winner_str = ws.cell(row=row, column=16).value
        home_away = ws.cell(row=row, column=7).value
        if winner_name is None or winner_str is None or home_away is None:
            continue
        m = GAME_PATTERN.match(str(winner_str))
        if not m:
            unmatched.append((row, winner_name, "no string match"))
            continue

        _, gtype, opp_dex, pf, pa, *_, elo_after = m.groups()
        opp_dex, pf, pa, elo_after = int(opp_dex), int(pf), int(pa), float(elo_after)
        winner_id = team_ids_by_name.get(winner_name)
        if winner_id is None:
            unmatched.append((row, winner_name, "no team_id match"))
            continue

        host_team_id = winner_id if str(home_away).strip().upper() == "H" else opp_dex
        host_region = team_regions.get(host_team_id)
        if host_region is None:
            unmatched.append((row, winner_name, f"no region for host team {host_team_id}"))
            continue

        row_a = conn.execute("""
            SELECT game_id FROM games WHERE season = 9 AND team_a = ? AND team_b = ?
            AND pf_a = ? AND pa_a = ? AND elo_a_after = ?
        """, (winner_id, opp_dex, pf, pa, elo_after)).fetchone()
        row_b = None
        if not row_a:
            row_b = conn.execute("""
                SELECT game_id FROM games WHERE season = 9 AND team_b = ? AND team_a = ?
                AND pa_a = ? AND pf_a = ? AND elo_b_after = ?
            """, (winner_id, opp_dex, pf, pa, elo_after)).fetchone()

        game_id = (row_a or row_b)
        if game_id is None:
            unmatched.append((row, winner_name, "no matching game"))
            continue
        conn.execute("UPDATE games SET host_region = ? WHERE game_id = ?",
                     (host_region, game_id["game_id"]))
        matched += 1

    conn.commit()
    return matched, unmatched


def migrate_accolades(wb, conn):
    """Backfill teams.accolades from Rankings column 79 -- real text
    (championship history etc.), found unmigrated on a later audit."""
    ws = wb["Rankings"]
    updated = 0
    for row in range(2, 162):
        dex = ws.cell(row=row, column=1).value
        acc = ws.cell(row=row, column=79).value
        if dex is None or not acc or not str(acc).strip():
            continue
        conn.execute("UPDATE teams SET accolades = ? WHERE team_id = ?", (str(acc).strip(), dex))
        updated += 1
    conn.commit()
    return updated


def run_full_migration(workbook_path=None):
    """Run the complete Season 9 migration end to end: teams, games,
    walkovers, EX bonus, fatigue, host regions, RDS Cup tagging, and the
    corrected fatigue recompute. Safe to re-run from scratch (init_db is
    idempotent; re-migrating overwrites prior data for the same season)."""
    from deckfield_ratings import add_walkover, get_connection, recompute_all_fatigue_deltas

    wb = openpyxl.load_workbook(workbook_path or XLSM_PATH, data_only=True, keep_vba=True)

    init_db()
    seed_region_distances()

    n_teams = migrate_teams(wb)
    print(f"Migrated {n_teams} teams")

    records, walkovers, unmatched = parse_game_records(wb)
    print(f"Parsed {len(records)} team-perspective game records, "
          f"{len(walkovers)} walkovers, {len(unmatched)} unmatched cells")

    games, skipped = dedupe_games(records)
    print(f"Deduplicated to {len(games)} games ({skipped} incomplete pairs skipped)")

    migrate_games(games)
    print("Games inserted.")

    for w in walkovers:
        add_walkover(SEASON, w["round"], w["game_type"], w["dex"], w["result"])
    print(f"{len(walkovers)} walkovers inserted.")

    conn = get_connection()
    matched, unmatched = migrate_ex_bonus(wb, conn)
    print(f"EX bonus backfilled for {matched} games, {len(unmatched)} unmatched")
    if unmatched:
        print("First few unmatched:", unmatched[:5])

    updated_seasons, delta_count = migrate_fatigue(wb, conn)
    print(f"Fatigue: {updated_seasons} starting values set, {delta_count} deltas inserted")

    host_matched, host_unmatched = migrate_host_region(wb, conn)
    print(f"Host region backfilled for {host_matched} games, {len(host_unmatched)} unmatched")
    if host_unmatched:
        print("First few unmatched:", host_unmatched[:5])

    rds_tagged, rds_unmatched = tag_rds_cup_games(wb, conn)
    print(f"RDS Cup games tagged: {rds_tagged}, {len(rds_unmatched)} unmatched")
    if rds_unmatched:
        print("First few unmatched:", rds_unmatched[:5])

    acc_updated = migrate_accolades(wb, conn)
    print(f"Accolades backfilled for {acc_updated} teams")

    conn.execute("""
        DELETE FROM fatigue_deltas WHERE (team_id, season, round) NOT IN (
            SELECT team_a, season, round FROM games
            UNION
            SELECT team_b, season, round FROM games
        )
    """)
    conn.commit()
    conn.close()

    recomputed = recompute_all_fatigue_deltas(SEASON)
    print(f"Fatigue deltas recomputed from real host_region history: {recomputed}")


if __name__ == "__main__":
    run_full_migration()
